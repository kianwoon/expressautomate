"""Turn an uploaded file's bytes into rows, and nothing more.

This module is pure: no database, no network, no settings. It answers two
questions — "what kind of table is this?" and "what rows does it hold?" —
and leaves everything about what those rows *mean* (matching candidates,
recording history, writing to the database) to later tasks.

XLSX is a zip container, exactly like the DOCX files `app/services/cv/text.py`
already had to defend against a reverse zip bomb for: a member's declared
size is a claim the file makes about itself, not a fact. So a `.xlsx` never
reaches `openpyxl` as the bytes we were handed — it goes through
`bounded_archive` first, which inflates every member itself against a
budget and repacks the result to `ZIP_STORED`. `openpyxl` reads that repack
exactly as it would read the original; only the trust boundary moves.
"""

import csv
import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Literal

import openpyxl

from app.services.archive import archive_contains, bounded_archive

# The dict key under which a CSV's single sheet is returned, unless the
# caller names it explicitly via `sheet_name`. A CSV has no internal notion
# of sheet names — unlike XLSX, where `Candidates` and `History` are real
# sheet names taken from the workbook — so we need some fixed default rather
# than a name lifted out of the file itself.
_CSV_SHEET_NAME = "csv"


class TooManyRows(Exception):
    """Raised when a sheet has more data rows than the caller's `max_rows`."""

    pass


class UnreadableTable(Exception):
    """Raised when the bytes claim to be a table but cannot be parsed as one."""

    pass


@dataclass(frozen=True)
class RowWidthProblem:
    """One CSV row whose cell count did not match the header's.

    A wider row almost always means an unquoted delimiter — a comma inside a
    description — silently shifting every later column. The row is skipped
    and this is reported alongside the other row problems, so one bad row
    costs itself rather than the file (the same rule `rows.py` keeps). A
    *shorter* row is the ordinary case of trailing empty cells and stays
    allowed: the missing columns read as empty, which per-row parsers report
    as "no email or phone" rather than silently misreading anything.
    """

    sheet: str
    line: int
    reason: str


def _decode_csv(data: bytes) -> str:
    """The file's text, whether it was saved as UTF-8 or UTF-16.

    UTF-8 first, with the BOM eaten by `utf-8-sig`. Excel's "Unicode Text"
    export writes UTF-16 with a byte-order mark, and the `utf-16` codec
    honours that mark and strips it — but only when a mark is actually
    there. Without one the endianness is unknowable, so the file is refused
    rather than guessed at.
    """
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        try:
            return data.decode("utf-16")
        except UnicodeDecodeError as exc:
            raise UnreadableTable("csv is not valid utf-16") from exc
    raise UnreadableTable("csv is not valid utf-8")


def _delimiter_for(text: str) -> str:
    """The separator this CSV most likely uses, judged from its header line.

    A header line is almost never quoted, so counting raw `,`, `;` and tab
    characters on that one line is a reliable proxy — and it is the one
    place where an Excel-locale file (semicolon or tab instead of comma)
    can be told apart from a comma file without a flaky heuristic. Comma
    wins ties, keeping the conventional default for a single-column file.
    """
    header_line = text.splitlines()[0] if text.splitlines() else ""
    semis = header_line.count(";")
    tabs = header_line.count("\t")
    commas = header_line.count(",")
    if semis > commas and semis > tabs:
        return ";"
    if tabs > commas and tabs > semis:
        return "\t"
    return ","


def sniff_table(data: bytes) -> Literal["csv", "xlsx"] | None:
    """Identify a table format from bytes alone — never filename, never content type.

    XLSX only counts if the zip's central directory actually lists
    `xl/workbook.xml`. `PK` at the front proves nothing: every zip, from a
    DOCX to a JAR to an XLSX, starts with the same two bytes, so checking
    only the magic number would call a Word document a spreadsheet.

    CSV is the fallback once the bytes decode as text — there is no magic
    number to check, only the absence of one.
    """
    if archive_contains(data, "xl/workbook.xml"):
        return "xlsx"

    try:
        _decode_csv(data)
    except UnreadableTable:
        return None

    return "csv"


def _normalise_header(raw: str | None) -> str:
    return (raw or "").strip().lower()


def _normalise_cell(value: object) -> str:
    """Render any cell value as text, collapsing whitespace-only to empty.

    A cell containing only spaces is not meaningfully different from an
    empty cell to anything downstream that matches on names or emails, but
    `" " != ""` would make every later equality check carry that special
    case. We resolve it here, once.
    """
    if value is None:
        return ""
    text = str(value).strip()
    return text


def _check_no_duplicate_headers(header: list[str]) -> None:
    """Refuse a header where two columns normalise to the same key.

    `_rows_from_records` builds a dict keyed by normalised header, so a
    collision — `Email` and `email ` in the same sheet — would silently
    drop one column and keep the other, last-write-wins, with no error and
    no warning. On an import that writes to live candidate records, reading
    the wrong column silently is worse than refusing the file outright.
    """
    seen: set[str] = set()
    for name in header:
        if not name:
            continue
        if name in seen:
            raise UnreadableTable(f"duplicate header: {name!r}")
        seen.add(name)


def _rows_from_records(
    header: list[str],
    records: Iterable[list[object]],
    max_rows: int,
    *,
    strict_width: bool = False,
    width_problems: list[RowWidthProblem] | None = None,
    sheet: str = "",
) -> list[dict[str, str]]:
    """Turn header + data rows into dicts, enforcing `max_rows` as we go.

    `records` is consumed lazily, one row at a time, so the cap is checked
    *before* each row is materialised into a dict rather than after the
    whole sheet has already been read into memory — the caller is expected
    to hand us a genuine iterator (a csv reader, an openpyxl row iterator),
    never a pre-built list, or this buys nothing.

    With `strict_width` (CSV only), a row with *more* cells than the header
    is skipped and reported in `width_problems` rather than having its tail
    silently dropped — an unquoted comma has shifted the columns and the
    extra cell is the only trace of it. The sheet's own name travels in for
    the report, since a CSV's one sheet is standing in for one of the two
    named sheets. XLSX never sets the flag: its rows are uniformly padded
    and ragged widths there are an openpyxl quirk, not a corruption signal.
    """
    _check_no_duplicate_headers(header)
    rows: list[dict[str, str]] = []
    for index, record in enumerate(records):
        if len(rows) >= max_rows:
            raise TooManyRows(f"sheet has more than {max_rows} rows")
        if strict_width and len(record) > len(header):
            # The CSV path is the only strict_width caller and always
            # supplies the list; the row is one problem, never a file error.
            assert width_problems is not None
            width_problems.append(
                RowWidthProblem(
                    sheet=sheet,
                    line=index + 2,
                    reason=(
                        f"this row has {len(record)} columns but the header has "
                        f"{len(header)} — a comma or quote in the data has probably "
                        "shifted the columns"
                    ),
                )
            )
            continue
        row = {
            header[i]: _normalise_cell(record[i])
            for i in range(min(len(header), len(record)))
            if header[i]
        }
        rows.append(row)
    return rows


def _read_csv(
    data: bytes, max_rows: int, sheet_name: str
) -> tuple[dict[str, list[dict[str, str]]], list[RowWidthProblem]]:
    # `_decode_csv` raises `UnreadableTable` with the specific reason when
    # the bytes are not a CSV we can read; it propagates as-is.
    text = _decode_csv(data)

    reader = csv.reader(io.StringIO(text), delimiter=_delimiter_for(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return {sheet_name: []}, []
    except csv.Error as exc:
        raise UnreadableTable("csv could not be parsed") from exc

    header = [_normalise_header(cell) for cell in header_row]

    def _records() -> Iterable[list[object]]:
        # A generator, not a list: `csv.Error` raised mid-stream must
        # surface at the row it happens on, and — the point of this whole
        # rewrite — nothing here is allowed to buffer the file into memory
        # before `_rows_from_records` gets a chance to enforce `max_rows`.
        try:
            yield from reader
        except csv.Error as exc:
            raise UnreadableTable("csv could not be parsed") from exc

    width_problems: list[RowWidthProblem] = []
    rows = _rows_from_records(
        header,
        _records(),
        max_rows,
        strict_width=True,
        width_problems=width_problems,
        sheet=sheet_name,
    )
    return {sheet_name: rows}, width_problems


def _read_xlsx(
    data: bytes, budget: int, max_rows: int
) -> tuple[dict[str, list[dict[str, str]]], list[RowWidthProblem]]:
    try:
        repacked = bounded_archive(data, budget=budget)
        workbook = openpyxl.load_workbook(repacked, read_only=True, data_only=True)
    except UnreadableTable:
        raise
    except Exception as exc:
        # Anything openpyxl or zipfile throws for a malformed workbook is a
        # library exception, not one of ours — the caller's contract is
        # that a corrupt file becomes `UnreadableTable`, never a leak of
        # whatever internal error openpyxl happened to raise this version.
        raise UnreadableTable("xlsx could not be parsed") from exc

    sheets: dict[str, list[dict[str, str]]] = {}
    try:
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            row_iter = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(row_iter)
            except StopIteration:
                sheets[name] = []
                continue
            header = [_normalise_header(cell) for cell in header_row]
            # Feed `row_iter` straight into `_rows_from_records` rather than
            # collecting it into a list first — the whole point of opening
            # this workbook with `read_only=True` is to stream rows one at
            # a time instead of holding the sheet in memory, and building
            # `records = [list(r) for r in row_iter]` here would silently
            # throw that away: the cap in `_rows_from_records` would then
            # be checked only after the entire sheet was already resident.
            records = (list(record) for record in row_iter)
            sheets[name] = _rows_from_records(header, records, max_rows)
    except (TooManyRows, UnreadableTable):
        # Both are already the right exception with the right message
        # (e.g. naming the duplicated header) — letting the `except
        # Exception` below repackage them would replace that message with
        # the generic "xlsx could not be parsed".
        raise
    except Exception as exc:
        raise UnreadableTable("xlsx could not be parsed") from exc
    finally:
        workbook.close()

    return sheets, []


def read_sheets(
    data: bytes,
    kind: str,
    *,
    budget: int,
    max_rows: int,
    sheet_name: str | None = None,
) -> tuple[dict[str, list[dict[str, str]]], list[RowWidthProblem]]:
    """Read a table's rows, keyed by sheet name then by lower-cased header.

    Returns the rows and the `RowWidthProblem`s the read skipped, so a
    caller can turn the second into the same error report the row parsers
    fill — a CSV row wider than its header is one row's problem, not the
    file's.

    `budget` bounds how much plaintext an XLSX's zip members may inflate to
    (see module docstring); `max_rows` bounds how many data rows any single
    sheet may hold before we give up rather than build an unbounded list in
    memory. Both are the caller's numbers — this module hardcodes neither.

    `sheet_name` only applies to `kind == "csv"`: a CSV has exactly one
    sheet and no internal name for it, but a caller may need to know
    whether that one sheet is standing in for "Candidates" or "History" —
    Task 7's API accepts a CSV as either. Naming it here, rather than
    handing back the fixed `"csv"` key and making every caller re-key the
    result, keeps that constant a private detail of this module instead of
    something every caller has to remember.
    """
    if kind == "csv":
        return _read_csv(data, max_rows, sheet_name or _CSV_SHEET_NAME)
    if kind == "xlsx":
        return _read_xlsx(data, budget, max_rows)
    raise UnreadableTable(f"unknown table kind: {kind!r}")
