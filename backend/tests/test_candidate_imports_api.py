# allow-hardcode: the header names, sheet names and SQL below are the fixed
# vocabulary of the import format, not a matching or scoring oracle.
"""Uploading a spreadsheet, watching it, reading its errors and undoing it.

Every test here is a boundary rather than a feature. This route takes bytes
from a browser and writes them to shared object storage on the strength of a
cookie, so the questions worth asking are adversarial: can a caller reach
another agency's import, can they hand us a PNG named `.xlsx`, and does a
file that was saved but never queued leave a recruiter watching a spinner
nobody is behind.
"""

import io
import uuid
import zipfile

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient
from sqlalchemy import text

from app.api import candidate_imports
from app.core.config import settings
from app.main import app
from app.models.candidate import CandidateImport
from app.services.imports.rows import (
    CANDIDATE_HEADERS,
    CANDIDATE_SHEET,
    HISTORY_HEADERS,
    HISTORY_SHEET,
)
from app.services.storage.r2 import InMemoryBodyStore
from tests.conftest import AdminSessionLocal
from tests.test_opportunities_api import sign_in  # the real session cookie, not a copy


async def _seed_agency() -> tuple[uuid.UUID, uuid.UUID]:
    tid, uid = uuid.uuid4(), uuid.uuid4()
    async with AdminSessionLocal() as s:
        await s.execute(
            text("INSERT INTO tenants (id, name, slug) VALUES (:i, :n, :n)"),
            {"i": tid, "n": f"agency-{tid.hex[:6]}"},
        )
        await s.execute(
            text("INSERT INTO users (id, tenant_id, email, role) VALUES (:i, :t, :e, 'owner')"),
            {"i": uid, "t": tid, "e": f"u{uid.hex[:6]}@agency.sg"},
        )
        await s.commit()
    return tid, uid


async def _drop_agency(tid: uuid.UUID) -> None:
    async with AdminSessionLocal() as s:
        for table in (
            "candidate_import_changes",
            "candidate_roles",
            "candidate_field_overrides",
            "candidate_skills",
            "candidates",
            "candidate_imports",
            "users",
        ):
            await s.execute(text(f"DELETE FROM {table} WHERE tenant_id = :t"), {"t": tid})
        await s.execute(text("DELETE FROM tenants WHERE id = :t"), {"t": tid})
        await s.commit()


@pytest.fixture
async def agency():
    tid, uid = await _seed_agency()
    yield tid, uid
    await _drop_agency(tid)


@pytest.fixture
async def other_agency():
    """A second agency, so "not yours" can be told apart from "not there"."""
    tid, uid = await _seed_agency()
    yield tid, uid
    await _drop_agency(tid)


@pytest.fixture
async def store():
    """The object store, swapped for the in-memory double.

    Through `app.dependency_overrides` rather than a patched global: the
    override is undone even when a test fails, so a failure cannot leave the
    next test writing to real R2.
    """
    double = InMemoryBodyStore()
    app.dependency_overrides[candidate_imports.body_store] = lambda: double
    yield double
    app.dependency_overrides.pop(candidate_imports.body_store, None)


@pytest.fixture
def queued(monkeypatch):
    """Every job the upload route tried to enqueue. Redis is never touched."""
    jobs: list[tuple[str, dict]] = []

    async def _enqueue(name: str, **kwargs) -> bool:
        jobs.append((name, kwargs))
        return True

    monkeypatch.setattr(candidate_imports, "enqueue", _enqueue)
    return jobs


@pytest.fixture
def enqueue_fails(monkeypatch):
    """Redis down: `enqueue` fails soft, exactly as `queue.py` does."""

    async def _enqueue(name: str, **kwargs) -> bool:
        return False

    monkeypatch.setattr(candidate_imports, "enqueue", _enqueue)


def _client_for(tid: uuid.UUID, uid: uuid.UUID) -> AsyncClient:
    http = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    sign_in(http, uid, tid)
    return http


def _csv_bytes() -> bytes:
    return b"full name,email\nJane Tan,jane@acme.sg\n"


def _xlsx_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    workbook.active.title = CANDIDATE_SHEET
    workbook.active.append(["full name", "email"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _png_named_xlsx() -> bytes:
    return b"\x89PNG\r\n\x1a\n" + b"\x00" * 64


async def _upload(http: AsyncClient, content: bytes, **kwargs):
    data = {"sheet": kwargs["sheet"]} if "sheet" in kwargs else None
    return await http.post(
        "/api/candidates/imports",
        files={"file": (kwargs.get("name", "roster.csv"), content, "text/csv")},
        data=data,
    )


async def _an_import(tenant_id: uuid.UUID, state: str, error_report_key: str | None = None):
    import_id = uuid.uuid4()
    async with AdminSessionLocal() as s:
        await s.execute(
            text(
                "INSERT INTO candidate_imports (id, tenant_id, filename, content_type,"
                " byte_size, object_key, state, error_report_key)"
                " VALUES (:i, :t, 'roster.csv', 'text/csv', 10, :k, :s, :e)"
            ),
            {
                "i": import_id,
                "t": tenant_id,
                "k": f"{tenant_id}/imports/{import_id}/candidates.csv",
                "s": state,
                "e": error_report_key,
            },
        )
        await s.commit()
    return import_id


async def _state(import_id: uuid.UUID) -> str:
    async with AdminSessionLocal() as s:
        return (
            await s.execute(
                text("SELECT state FROM candidate_imports WHERE id = :i"), {"i": import_id}
            )
        ).scalar_one()


# --- Upload -----------------------------------------------------------------


async def test_upload_stores_enqueues_and_returns_202(agency, store, queued):
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _csv_bytes())

    assert response.status_code == 202, response.text
    body = response.json()
    assert body["state"] == CandidateImport.PENDING
    assert body["filename"] == "roster.csv"

    # The acceptance criterion this route exists for: without the enqueue an
    # uploaded roster waits for whenever `rescan_stuck` next runs.
    assert queued == [
        ("run_candidate_import", {"tenant_id": str(tid), "import_id": body["id"]})
    ]
    assert store.binary_objects  # the bytes really were written


async def test_a_csv_records_which_sheet_it_is_in_the_key(agency, store, queued):
    """A CSV has one nameless sheet, so the route is the only place that
    knows whether it is a roster or a career history."""
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _csv_bytes(), sheet=HISTORY_SHEET)

    assert response.status_code == 202, response.text
    key = next(iter(store.binary_objects))
    assert key.startswith(f"{tid}/imports/")
    assert key.endswith("/history.csv")


async def test_an_xlsx_needs_no_sheet_because_it_names_its_own(agency, store, queued):
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _xlsx_bytes(), name="roster.xlsx")

    assert response.status_code == 202, response.text
    assert next(iter(store.binary_objects)).endswith("/workbook.xlsx")


async def test_an_unknown_sheet_name_is_refused(agency, store, queued):
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _csv_bytes(), sheet="Sheet1")

    assert response.status_code == 422, response.text
    assert not store.binary_objects


async def test_an_oversized_file_is_a_413(agency, store, queued, monkeypatch):
    monkeypatch.setattr(settings, "IMPORT_MAX_UPLOAD_BYTES", 8)
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _csv_bytes())

    assert response.status_code == 413, response.text
    assert "8 byte limit" in response.json()["detail"]
    # Nothing stored and nothing queued: the refusal happens before either.
    assert not store.binary_objects
    assert queued == []


async def test_a_png_named_xlsx_is_a_415(agency, store, queued):
    """The bytes decide, never the extension and never the Content-Type."""
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _png_named_xlsx(), name="roster.xlsx")

    assert response.status_code == 415, response.text
    assert not store.binary_objects
    assert queued == []


async def test_a_docx_is_not_a_spreadsheet(agency, store, queued):
    """`PK` proves nothing: every zip starts with it. Only `xl/workbook.xml`
    makes an archive a spreadsheet."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("word/document.xml", "<w:document/>")

    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, buffer.getvalue(), name="roster.xlsx")

    assert response.status_code == 415, response.text


async def test_a_failed_enqueue_leaves_the_import_failed_not_pending(
    agency, store, enqueue_fails
):
    """The hole this exists to close: `enqueue` returns a bool and never
    raises, so a silent Redis outage would otherwise leave the row `pending`
    for ever and the recruiter watching a spinner nobody is behind."""
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await _upload(http, _csv_bytes())

    assert response.status_code == 202, response.text
    body = response.json()
    assert body["state"] == CandidateImport.FAILED
    assert body["has_errors"] is True
    assert await _state(uuid.UUID(body["id"])) == CandidateImport.FAILED

    # The reason is readable, because an import has no error column.
    report, content_type = store.binary_objects[f"{tid}/imports/{body['id']}/errors.txt"]
    assert b"could not be queued" in report
    assert content_type.startswith("text/plain")


# --- Listing ----------------------------------------------------------------


async def test_the_list_shows_only_this_agencys_imports(agency, other_agency):
    tid, uid = agency
    other_tid, _other_uid = other_agency
    mine = await _an_import(tid, CandidateImport.DONE)
    theirs = await _an_import(other_tid, CandidateImport.DONE)

    async with _client_for(tid, uid) as http:
        body = (await http.get("/api/candidates/imports")).json()

    ids = {row["id"] for row in body}
    assert str(mine) in ids
    assert str(theirs) not in ids


# --- The error report -------------------------------------------------------


async def test_errors_returns_a_short_lived_url(agency, store):
    tid, uid = agency
    import_id = await _an_import(
        tid, CandidateImport.DONE, error_report_key=f"{tid}/imports/x/errors.txt"
    )

    async with _client_for(tid, uid) as http:
        response = await http.get(f"/api/candidates/imports/{import_id}/errors")

    assert response.status_code == 200, response.text
    assert response.json()["expires_in"] == settings.IMPORT_PRESIGNED_URL_TTL_SECONDS


async def test_errors_on_a_clean_import_is_a_404(agency, store):
    tid, uid = agency
    import_id = await _an_import(tid, CandidateImport.DONE)

    async with _client_for(tid, uid) as http:
        response = await http.get(f"/api/candidates/imports/{import_id}/errors")

    assert response.status_code == 404, response.text


async def test_another_agencys_error_report_is_a_404_not_a_403(agency, other_agency, store):
    tid, uid = agency
    other_tid, _other_uid = other_agency
    theirs = await _an_import(
        other_tid, CandidateImport.DONE, error_report_key=f"{other_tid}/imports/x/errors.txt"
    )

    async with _client_for(tid, uid) as http:
        response = await http.get(f"/api/candidates/imports/{theirs}/errors")

    assert response.status_code == 404, response.text


# --- Undo -------------------------------------------------------------------


async def test_another_agencys_import_cannot_be_undone(agency, other_agency):
    tid, uid = agency
    other_tid, _other_uid = other_agency
    theirs = await _an_import(other_tid, CandidateImport.DONE)

    async with _client_for(tid, uid) as http:
        response = await http.post(f"/api/candidates/imports/{theirs}/undo")

    assert response.status_code == 404, response.text
    # And it really was left alone.
    assert await _state(theirs) == CandidateImport.DONE


async def test_undo_refuses_while_the_import_is_still_running(agency):
    """409 rather than 404: the import is the caller's own, and the state is
    the objection — undoing now would race the run still writing."""
    tid, uid = agency
    import_id = await _an_import(tid, CandidateImport.PARSING)

    async with _client_for(tid, uid) as http:
        response = await http.post(f"/api/candidates/imports/{import_id}/undo")

    assert response.status_code == 409, response.text
    assert await _state(import_id) == CandidateImport.PARSING


async def test_undo_is_idempotent(agency):
    tid, uid = agency
    import_id = await _an_import(tid, CandidateImport.DONE)

    async with _client_for(tid, uid) as http:
        first = await http.post(f"/api/candidates/imports/{import_id}/undo")
        second = await http.post(f"/api/candidates/imports/{import_id}/undo")

    assert first.status_code == 200, first.text
    assert first.json()["already_undone"] is False
    assert second.status_code == 200, second.text
    assert second.json()["already_undone"] is True
    assert await _state(import_id) == CandidateImport.UNDONE


# --- The template -----------------------------------------------------------


async def test_the_template_carries_exactly_the_headers_the_parser_reads(agency):
    """A template naming a column `rows.py` does not read is worse than no
    template: the recruiter fills it in and the value disappears."""
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await http.get("/api/candidates/imports/template")

    assert response.status_code == 200, response.text
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [CANDIDATE_SHEET, HISTORY_SHEET]
    assert [c.value for c in workbook[CANDIDATE_SHEET][1]] == list(CANDIDATE_HEADERS)
    assert [c.value for c in workbook[HISTORY_SHEET][1]] == list(HISTORY_HEADERS)


async def test_the_template_path_is_not_read_as_an_import_id(agency):
    """Declared before `{import_id}`, so `template` never becomes a 422."""
    tid, uid = agency
    async with _client_for(tid, uid) as http:
        response = await http.get("/api/candidates/imports/template")
    assert response.status_code == 200
